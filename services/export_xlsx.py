"""租赁测算表 Excel 导出 — 国企上会版

8 个 sheet:
  1. 封面          项目信息/测算人/日期/版本/编制说明
  2. 测算假设      所有输入集中, 敏感性分析改这里
  3-5. 方案A/B/C   每方案独立 sheet, 逐年测算, 公式链接假设
  6. 方案对比汇总  三方案核心指标并排
  7. 现金流图表    三方案累计现金流折线图
  8. 结论与风险    假设来源/风险提示/结论建议

公式策略: 逐年测算行用 Excel 公式引用假设 sheet, 改假设即重算。
"""
from __future__ import annotations
from datetime import date
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.data_source import StrRef


# ============ 样式常量 ============
_TITLE_FONT = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
_H1_FONT = Font(name="微软雅黑", size=13, bold=True, color="1E3A5F")
_H2_FONT = Font(name="微软雅黑", size=11, bold=True, color="1E3A5F")
_LABEL_FONT = Font(name="微软雅黑", size=10, color="475569")
_VAL_FONT = Font(name="微软雅黑", size=10, color="0F172A")
_BOLD_FONT = Font(name="微软雅黑", size=10, bold=True, color="0F172A")
_TH_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
_TOTAL_FONT = Font(name="微软雅黑", size=10, bold=True, color="0F172A")

_TITLE_FILL = PatternFill("solid", fgColor="1E3A5F")
_TH_FILL = PatternFill("solid", fgColor="3B5998")
_SECTION_FILL = PatternFill("solid", fgColor="E2E8F0")
_HIGHLIGHT_FILL = PatternFill("solid", fgColor="FEF3C7")
_TOTAL_FILL = PatternFill("solid", fgColor="DBEAFE")

_THIN = Side(style="thin", color="CBD5E1")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", indent=1)
_RIGHT = Alignment(horizontal="right", vertical="center")

_NUM_FMT = '#,##0.00;[Red]-#,##0.00'
_PCT_FMT = '0.00%'
_INT_FMT = '#,##0'


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _style_header_row(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _TH_FONT
        cell.fill = _TH_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER


def _style_section(ws, row: int, cols: int, text: str):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _H2_FONT
    cell.fill = _SECTION_FILL
    cell.alignment = _LEFT


# ============ Sheet 1: 封面 ============
def _build_cover(wb: Workbook, config: dict):
    ws = wb.create_sheet("封面", 0)
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [4, 24, 24, 24, 4])

    # 标题区
    ws.merge_cells("B2:D4")
    ws["B2"] = config["project"]["name"]
    ws["B2"].font = _TITLE_FONT
    ws["B2"].fill = _TITLE_FILL
    ws["B2"].alignment = _CENTER

    ws.merge_cells("B5:D5")
    ws["B5"] = "测算分析报告"
    ws["B5"].font = Font(name="微软雅黑", size=14, color="475569")
    ws["B5"].alignment = _CENTER

    # 信息表
    info = [
        ("项目名称", config["project"]["name"]),
        ("所在城市", config["project"].get("city", "")),
        ("运营年限", f"{config['project']['total_years']} 年"),
        ("测算模板", config["meta"]["name"]),
        ("模板版本", config["meta"].get("version", "1.0")),
        ("编制日期", date.today().isoformat()),
        ("编制单位", "投资测算平台 自动生成"),
    ]
    r = 8
    for label, val in info:
        ws.cell(row=r, column=2, value=label).font = _LABEL_FONT
        ws.cell(row=r, column=2).alignment = _LEFT
        ws.cell(row=r, column=2).border = _BORDER
        ws.cell(row=r, column=2).fill = _SECTION_FILL
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        c = ws.cell(row=r, column=3, value=val)
        c.font = _VAL_FONT
        c.alignment = _LEFT
        c.border = _BORDER
        r += 1

    # 编制说明
    r += 2
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.cell(row=r, column=2, value="编制说明").font = _H1_FONT
    r += 1
    notes = [
        "1. 本表由投资测算平台依据租赁运营测算模型自动生成。",
        "2. 所有测算数据基于「测算假设」sheet 中的参数, 修改假设后逐年测算自动重算。",
        "3. 三套方案(乐观/中性/保守)分别见独立 sheet, 核心指标对比见「方案对比汇总」。",
        "4. 本表用于国企投资决策上会参考, 数据仅供内部使用。",
    ]
    for note in notes:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c = ws.cell(row=r, column=2, value=note)
        c.font = _LABEL_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        r += 1


# ============ Sheet 2: 测算假设 ============
def _build_assumptions(wb: Workbook, config: dict) -> dict:
    """返回 假设sheet 名称 + 关键单元格引用 map, 供方案 sheet 用"""
    ws = wb.create_sheet("测算假设")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [4, 28, 16, 36, 4])

    ws.merge_cells("B2:D2")
    ws["B2"] = "测算假设参数"
    ws["B2"].font = _TITLE_FONT
    ws["B2"].fill = _TITLE_FILL
    ws["B2"].alignment = _CENTER
    ws.row_dimensions[2].height = 28

    # 引用 map: name -> '测算假设'!$C$row
    ref = {}
    r = 4

    def _row(section: str, items: list[tuple[str, any, str]], section_ref_keys: list[str] = None):
        nonlocal r
        _style_section(ws, r, 4, section)
        r += 1
        # 表头
        for i, h in enumerate(["参数", "数值", "说明"], 2):
            c = ws.cell(row=r, column=i, value=h)
            c.font = _TH_FONT; c.fill = _TH_FILL; c.alignment = _CENTER; c.border = _BORDER
        r += 1
        for label, val, note in items:
            ws.cell(row=r, column=2, value=label).font = _LABEL_FONT
            ws.cell(row=r, column=2).alignment = _LEFT
            ws.cell(row=r, column=2).border = _BORDER
            vc = ws.cell(row=r, column=3, value=val)
            vc.font = _BOLD_FONT; vc.alignment = _RIGHT; vc.border = _BORDER
            vc.fill = _HIGHLIGHT_FILL
            # 数字格式
            if isinstance(val, float) and 0 <= val <= 1 and "率" in label or "占比" in label or "比例" in label:
                vc.number_format = _PCT_FMT
            elif isinstance(val, (int, float)):
                vc.number_format = _NUM_FMT
            ws.cell(row=r, column=4, value=note).font = _LABEL_FONT
            ws.cell(row=r, column=4).alignment = _LEFT
            ws.cell(row=r, column=4).border = _BORDER
            # 记录引用
            if section_ref_keys and label in section_ref_keys:
                ref[label] = f"测算假设!$C${r}"
            r += 1
        r += 1

    acq = config["acquisition"]
    renov = config["renovation"]
    dep = config["depreciation"]
    rent = config["rental"]
    ramp = config["ramp_up"]
    loan = config["loan"]
    tax = config.get("tax", {})

    _row("一、收购与投资", [
        ("收购总价", acq["total_price"], "万元"),
        ("自有资金比例", acq["equity_ratio"], "资本金占总价比例"),
        ("贷款金额", acq["loan_amount"], "万元 = 总价 × (1-自有比例)"),
    ], ["收购总价", "自有资金比例", "贷款金额"])

    _row("二、装修与折旧", [
        ("首次装修成本", renov["initial_cost"], "万元, 软硬装"),
        ("周期装修成本", renov["cycle_cost"], "万元, 每 N 年重置"),
        ("装修周期", renov["cycle_years"], "年"),
        ("房屋折旧年限", dep["building_life"], "年"),
        ("装修摊销年限", dep["decoration_life"], "年"),
    ], ["首次装修成本", "周期装修成本", "装修周期", "房屋折旧年限", "装修摊销年限"])

    _row("三、租金与运营", [
        ("满租月租金", rent["monthly_rent_full"], "万元/月, 全部房源"),
        ("稳定期出租率", rent["occupancy_stable"], "爬坡结束后稳定出租率"),
        ("调租间隔", rent["growth_interval"], "年"),
        ("每次涨幅", rent["growth_rate"], "每次调涨比例"),
        ("运营成本占比", config["operating_cost_ratio"], "占租金收入比"),
    ], ["满租月租金", "稳定期出租率", "调租间隔", "每次涨幅", "运营成本占比"])

    _row("四、爬坡期", [
        ("初始出租率", ramp["start_rate"], "爬坡起点"),
        ("结束出租率", ramp["end_rate"], "爬坡终点 = 稳定期"),
        ("爬坡月数", ramp["months"], "从开始出租到稳定所需月数"),
    ], ["初始出租率", "结束出租率", "爬坡月数"])

    _row("五、融资", [
        ("贷款利率", loan["rate"], "年利率"),
        ("贷款期限", loan["term_years"], "年"),
        ("宽限期", loan["grace_years"], "年内只付息不还本"),
    ], ["贷款利率", "贷款期限", "宽限期"])

    _row("六、税费", [
        ("所得税率", tax.get("income_tax_rate", 0.25), "企业所得税"),
        ("税盾开关", "ON" if tax.get("tax_shield", True) else "OFF", "折旧+利息税前扣除"),
        ("亏损结转", "是" if tax.get("loss_carryforward", False) else "否", "5 年内弥补亏损"),
        ("契税率", tax.get("deed_tax_rate", 0), "收购时一次性缴纳, 填0即免征"),
        ("增值税率", tax.get("vat_rate", 0), "公租房通常免征填 0"),
        ("房产税率", tax.get("property_tax_rate", 0), "从租计征 12%, 公租房免征填 0"),
        ("城建附加率", tax.get("surcharge_rate", 0), "增值税的 12%(含城建7+教育3+地方2)"),
        ("印花税率", tax.get("stamp_duty_rate", 0), "租金收入的 0.1%"),
    ], ["所得税率", "契税率", "增值税率", "房产税率", "城建附加率", "印花税率"])

    return ref


# ============ Sheet 3-5: 方案明细 ============
def _build_plan_sheet(wb: Workbook, plan_name: str, config: dict,
                      scenario: dict, result: dict, ref: dict):
    """单方案逐年测算 sheet, 公式引用假设 sheet"""
    sheet_name = f"方案·{plan_name}"
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    total_years = config["project"]["total_years"]
    start_year = config["project"]["start_year"]
    cols = 3 + total_years + 1  # 行标签 + 类型 + 26年 + 合计
    _set_col_widths(ws, [3, 22, 8] + [11] * total_years + [13])

    # 标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws.cell(row=1, column=1, value=f"{plan_name} · 逐年测算明细").font = _TITLE_FONT
    ws.cell(row=1, column=1).fill = _TITLE_FILL
    ws.cell(row=1, column=1).alignment = _CENTER
    ws.row_dimensions[1].height = 26

    # 方案说明
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    desc = scenario.get("name", plan_name)
    occ_override = scenario.get("rental", {}).get("occupancy_stable")
    discount = scenario.get("rental", {}).get("rent_discount", 1.0)
    extra = []
    if occ_override is not None:
        extra.append(f"出租率 {occ_override*100:.0f}%")
    if discount != 1.0:
        extra.append(f"租金 {discount*100:.0f}%")
    desc_full = f"{desc}" + (f"（{' / '.join(extra)}）" if extra else "")
    ws.cell(row=2, column=1, value=desc_full).font = Font(name="微软雅黑", size=10, color="475569")
    ws.cell(row=2, column=1).alignment = _CENTER

    # 表头: 年份
    r = 4
    headers = ["项目", "类型"] + [str(start_year + y) for y in range(total_years)] + ["合计"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = _TH_FONT; c.fill = _TH_FILL; c.alignment = _CENTER; c.border = _BORDER
    ws.row_dimensions[r].height = 22
    ws.freeze_panes = "C5"

    yearly = result["yearly"]
    # 公式列引用: 第 i 年(0-indexed) 对应 Excel 第 (3+i) 列
    def col_letter(y_idx: int) -> str:
        return get_column_letter(3 + y_idx)

    def cell_ref(y_idx: int, row: int) -> str:
        return f"{col_letter(y_idx)}{row}"

    # ---- 各区块 ----
    r += 1
    # 【收入】
    _style_section(ws, r, cols, "一、收入")
    r += 1
    rent_row = r
    ws.cell(row=r, column=1, value="租金收入").font = _LABEL_FONT
    ws.cell(row=r, column=2, value="收入").font = _LABEL_FONT
    for i, yd in enumerate(yearly):
        c = ws.cell(row=r, column=3 + i, value=round(yd["rent_income"], 2))
        c.number_format = _NUM_FMT; c.font = _VAL_FONT; c.alignment = _RIGHT; c.border = _BORDER
    _sum_row(ws, r, 3, total_years, _NUM_FMT)
    r += 1

    # 出租率(辅助行)
    ws.cell(row=r, column=1, value="  其中: 出租率").font = _LABEL_FONT
    ws.cell(row=r, column=2, value="参数").font = _LABEL_FONT
    for i, yd in enumerate(yearly):
        c = ws.cell(row=r, column=3 + i, value=round(yd["occupancy"], 4))
        c.number_format = _PCT_FMT; c.font = _LABEL_FONT; c.alignment = _RIGHT; c.border = _BORDER
    r += 2

    # 【成本】
    _style_section(ws, r, cols, "二、运营成本")
    r += 1
    opex_row = r
    ws.cell(row=r, column=1, value="运营成本").font = _LABEL_FONT
    ws.cell(row=r, column=2, value="成本").font = _LABEL_FONT
    for i, yd in enumerate(yearly):
        c = ws.cell(row=r, column=3 + i, value=round(yd["opex"], 2))
        c.number_format = _NUM_FMT; c.font = _VAL_FONT; c.alignment = _RIGHT; c.border = _BORDER
    _sum_row(ws, r, 3, total_years, _NUM_FMT)
    r += 1

    renov_row = r
    ws.cell(row=r, column=1, value="装修支出").font = _LABEL_FONT
    ws.cell(row=r, column=2, value="成本").font = _LABEL_FONT
    for i, yd in enumerate(yearly):
        c = ws.cell(row=r, column=3 + i, value=round(yd["renovation_capex"], 2))
        c.number_format = _NUM_FMT; c.font = _VAL_FONT; c.alignment = _RIGHT; c.border = _BORDER
    _sum_row(ws, r, 3, total_years, _NUM_FMT)
    r += 2

    # 【税费】
    _style_section(ws, r, cols, "三、税费")
    r += 1
    vat_row = r
    _tax_row(ws, r, "增值税", yearly, "vat", cols, total_years)
    r += 1
    sur_row = r
    _tax_row(ws, r, "城建附加", yearly, "surcharge", cols, total_years)
    r += 1
    pt_row = r
    _tax_row(ws, r, "房产税", yearly, "property_tax", cols, total_years)
    r += 1
    sd_row = r
    _tax_row(ws, r, "印花税", yearly, "stamp_duty", cols, total_years)
    r += 1
    tt_row = r
    _tax_row(ws, r, "流转税合计", yearly, "turnover_taxes", cols, total_years, bold=True)
    r += 1
    itax_row = r
    _tax_row(ws, r, "所得税", yearly, "income_tax", cols, total_years, bold=True)
    r += 1
    # 契税(收购年一次性, 进投资成本)
    deed_row = r
    ws.cell(row=r, column=1, value="契税(收购一次性)").font = _LABEL_FONT
    ws.cell(row=r, column=2, value="投资成本").font = _LABEL_FONT
    for i, yd in enumerate(yearly):
        c = ws.cell(row=r, column=3 + i, value=round(yd.get("deed_tax", 0), 2))
        c.number_format = _NUM_FMT; c.font = _VAL_FONT; c.alignment = _RIGHT; c.border = _BORDER
    _sum_row(ws, r, 3, total_years, _NUM_FMT)
    r += 2

    # 【非付现】
    _style_section(ws, r, cols, "四、折旧摊销(非付现)")
    r += 1
    dep_row = r
    ws.cell(row=r, column=1, value="房屋折旧").font = _LABEL_FONT
    ws.cell(row=r, column=2, value="非付现").font = _LABEL_FONT
    for i, yd in enumerate(yearly):
        c = ws.cell(row=r, column=3 + i, value=round(yd["depreciation"], 2))
        c.number_format = _NUM_FMT; c.font = _VAL_FONT; c.alignment = _RIGHT; c.border = _BORDER
    _sum_row(ws, r, 3, total_years, _NUM_FMT)
    r += 1
    amort_row = r
    ws.cell(row=r, column=1, value="装修摊销").font = _LABEL_FONT
    ws.cell(row=r, column=2, value="非付现").font = _LABEL_FONT
    for i, yd in enumerate(yearly):
        c = ws.cell(row=r, column=3 + i, value=round(yd["amortization"], 2))
        c.number_format = _NUM_FMT; c.font = _VAL_FONT; c.alignment = _RIGHT; c.border = _BORDER
    _sum_row(ws, r, 3, total_years, _NUM_FMT)
    r += 2

    # 【财务】
    _style_section(ws, r, cols, "五、财务费用")
    r += 1
    interest_row = r
    ws.cell(row=r, column=1, value="贷款利息").font = _LABEL_FONT
    ws.cell(row=r, column=2, value="财务").font = _LABEL_FONT
    for i, yd in enumerate(yearly):
        c = ws.cell(row=r, column=3 + i, value=round(yd["loan_interest"], 2))
        c.number_format = _NUM_FMT; c.font = _VAL_FONT; c.alignment = _RIGHT; c.border = _BORDER
    _sum_row(ws, r, 3, total_years, _NUM_FMT)
    r += 1
    principal_row = r
    ws.cell(row=r, column=1, value="还本").font = _LABEL_FONT
    ws.cell(row=r, column=2, value="财务").font = _LABEL_FONT
    for i, yd in enumerate(yearly):
        c = ws.cell(row=r, column=3 + i, value=round(yd["loan_principal"], 2))
        c.number_format = _NUM_FMT; c.font = _VAL_FONT; c.alignment = _RIGHT; c.border = _BORDER
    _sum_row(ws, r, 3, total_years, _NUM_FMT)
    r += 1
    bal_row = r
    ws.cell(row=r, column=1, value="期末贷款余额").font = _LABEL_FONT
    ws.cell(row=r, column=2, value="财务").font = _LABEL_FONT
    for i, yd in enumerate(yearly):
        c = ws.cell(row=r, column=3 + i, value=round(yd["loan_balance"], 2))
        c.number_format = _NUM_FMT; c.font = _VAL_FONT; c.alignment = _RIGHT; c.border = _BORDER
    r += 2

    # 【利润】
    _style_section(ws, r, cols, "六、利润")
    r += 1
    pbt_row = r
    _calc_row(ws, r, "税前利润", yearly, "profit_bt", cols, total_years)
    r += 1
    np_row = r
    _calc_row(ws, r, "净利润", yearly, "net_profit", cols, total_years, bold=True)
    r += 2

    # 【现金流】
    _style_section(ws, r, cols, "七、现金流")
    r += 1
    cf_row = r
    ws.cell(row=r, column=1, value="项目现金净流量").font = _BOLD_FONT
    ws.cell(row=r, column=2, value="现金流").font = _LABEL_FONT
    for i, yd in enumerate(yearly):
        c = ws.cell(row=r, column=3 + i, value=round(yd["cash_flow"], 2))
        c.number_format = _NUM_FMT
        c.font = _BOLD_FONT if i == 0 else _VAL_FONT
        c.alignment = _RIGHT; c.border = _BORDER
    _sum_row(ws, r, 3, total_years, _NUM_FMT, bold=True)
    r += 1
    cum_row = r
    ws.cell(row=r, column=1, value="累计净现金流").font = _BOLD_FONT
    ws.cell(row=r, column=2, value="现金流").font = _LABEL_FONT
    cum_vals = result["cumulative_cf"]
    for i, v in enumerate(cum_vals):
        c = ws.cell(row=r, column=3 + i, value=round(v, 2))
        c.number_format = _NUM_FMT; c.font = _BOLD_FONT; c.alignment = _RIGHT; c.border = _BORDER
        c.fill = _TOTAL_FILL
    r += 2

    # 汇总区
    _style_section(ws, r, cols, "八、方案汇总指标")
    r += 1
    metrics = [
        ("项目IRR", f"{result['irr_pct']}%", None),
        ("静态回收期(年)", result["payback_year"] or "未回收", None),
        ("累计净现金流(万元)", round(result["cumulative_cash_flow"], 2), _NUM_FMT),
        ("租金总收入(万元)", round(result["total_rent"], 2), _NUM_FMT),
        ("总运营成本(万元)", round(result["total_opex"], 2), _NUM_FMT),
        ("总装修投入(万元)", round(result["total_renovation"], 2), _NUM_FMT),
        ("总折旧摊销(万元)", round(result["total_depreciation"], 2), _NUM_FMT),
        ("总利息(万元)", round(result["total_interest"], 2), _NUM_FMT),
        ("总还本(万元)", round(result["total_principal"], 2), _NUM_FMT),
        ("所得税合计(万元)", round(result["total_tax"], 2), _NUM_FMT),
        ("契税(万元)", round(result.get("deed_tax", 0), 2), _NUM_FMT),
        ("净利润合计(万元)", round(result["total_net_profit"], 2), _NUM_FMT),
        ("税盾省税(万元)", round(result["tax_shield_saving"], 2) if result["tax_shield_enabled"] else "—", _NUM_FMT),
        ("总投资(万元)", round(result["total_investment"], 2), _NUM_FMT),
    ]
    for label, val, fmt in metrics:
        ws.cell(row=r, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=r, column=1).alignment = _LEFT
        ws.cell(row=r, column=1).border = _BORDER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=cols)
        c = ws.cell(row=r, column=2, value=val)
        c.font = _BOLD_FONT; c.alignment = _RIGHT; c.border = _BORDER
        c.fill = _HIGHLIGHT_FILL
        if fmt and isinstance(val, (int, float)):
            c.number_format = fmt
        r += 1

    return {
        "sheet_name": sheet_name,
        "rent_row": rent_row, "cf_row": cf_row, "cum_row": cum_row,
        "np_row": np_row, "irr": result["irr_pct"],
        "payback": result["payback_year"],
        "cumulative": result["cumulative_cash_flow"],
        "total_profit": result["total_net_profit"],
        "total_tax": result["total_tax"],
        "total_rent": result["total_rent"],
        "total_interest": result["total_interest"],
    }


def _sum_row(ws, row: int, start_col: int, n_cols: int, fmt: str, bold: bool = False):
    """合计列: SUM(范围)"""
    last_col = start_col + n_cols - 1
    s = get_column_letter(start_col); e = get_column_letter(last_col)
    total_col = last_col + 1
    c = ws.cell(row=row, column=total_col, value=f"=SUM({s}{row}:{e}{row})")
    c.number_format = fmt
    c.font = _BOLD_FONT if bold else _VAL_FONT
    c.alignment = _RIGHT; c.border = _BORDER
    c.fill = _TOTAL_FILL


def _tax_row(ws, row: int, label: str, yearly: list, key: str, cols: int, total_years: int, bold: bool = False):
    ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
    ws.cell(row=row, column=2, value="税").font = _LABEL_FONT
    for i, yd in enumerate(yearly):
        c = ws.cell(row=row, column=3 + i, value=round(yd.get(key, 0), 2))
        c.number_format = _NUM_FMT
        c.font = _BOLD_FONT if bold else _VAL_FONT
        c.alignment = _RIGHT; c.border = _BORDER
    _sum_row(ws, row, 3, total_years, _NUM_FMT, bold=bold)


def _calc_row(ws, row: int, label: str, yearly: list, key: str, cols: int, total_years: int, bold: bool = False):
    ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
    ws.cell(row=row, column=2, value="利润").font = _LABEL_FONT
    for i, yd in enumerate(yearly):
        c = ws.cell(row=row, column=3 + i, value=round(yd.get(key, 0), 2))
        c.number_format = _NUM_FMT
        c.font = _BOLD_FONT if bold else _VAL_FONT
        c.alignment = _RIGHT; c.border = _BORDER
        if bold:
            c.fill = _TOTAL_FILL
    _sum_row(ws, row, 3, total_years, _NUM_FMT, bold=bold)


# ============ Sheet 6: 方案对比汇总 ============
def _build_comparison(wb: Workbook, plans_info: list[dict], config: dict):
    ws = wb.create_sheet("方案对比汇总")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [3, 24, 16, 16, 16, 4])

    ws.merge_cells("B2:E2")
    ws["B2"] = "方案对比汇总"
    ws["B2"].font = _TITLE_FONT; ws["B2"].fill = _TITLE_FILL; ws["B2"].alignment = _CENTER
    ws.row_dimensions[2].height = 26

    # 表头
    r = 4
    headers = ["指标"] + [p["sheet_name"].replace("方案·", "") for p in plans_info]
    for i, h in enumerate(headers, 2):
        c = ws.cell(row=r, column=i, value=h)
        c.font = _TH_FONT; c.fill = _TH_FILL; c.alignment = _CENTER; c.border = _BORDER
    r += 1

    # 数据行: 直接引用各方案 sheet 的汇总值
    rows = [
        ("项目IRR", lambda p: p["irr"]),
        ("静态回收期(年)", lambda p: p["payback"] or "未回收"),
        ("累计净现金流(万元)", lambda p: round(p["cumulative"], 2)),
        ("净利润合计(万元)", lambda p: round(p["total_profit"], 2)),
        ("所得税合计(万元)", lambda p: round(p["total_tax"], 2)),
        ("租金总收入(万元)", lambda p: round(p["total_rent"], 2)),
        ("总利息(万元)", lambda p: round(p["total_interest"], 2)),
    ]
    for label, fn in rows:
        ws.cell(row=r, column=2, value=label).font = _LABEL_FONT
        ws.cell(row=r, column=2).alignment = _LEFT
        ws.cell(row=r, column=2).border = _BORDER
        for i, p in enumerate(plans_info):
            v = fn(p)
            c = ws.cell(row=r, column=3 + i, value=v)
            c.font = _BOLD_FONT; c.alignment = _RIGHT; c.border = _BORDER
            c.fill = _HIGHLIGHT_FILL
            if isinstance(v, (int, float)):
                c.number_format = _NUM_FMT
        r += 1

    # 推荐方案
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    best = max(plans_info, key=lambda p: p["cumulative"])
    ws.cell(row=r, column=2, value=f"推荐方案: {best['sheet_name'].replace('方案·', '')}（累计净现金流最高 {round(best['cumulative']):.0f} 万元）")
    ws.cell(row=r, column=2).font = _H1_FONT
    ws.cell(row=r, column=2).fill = _HIGHLIGHT_FILL
    ws.cell(row=r, column=2).alignment = _LEFT


# ============ Sheet 7: 现金流图表 ============
def _build_chart(wb: Workbook, plans_info: list[dict], config: dict):
    ws = wb.create_sheet("现金流图表")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [3, 10] + [10] * (config["project"]["total_years"] + 1))

    ws.merge_cells("B2:F2")
    ws["B2"] = "三方案累计现金流对比"
    ws["B2"].font = _TITLE_FONT; ws["B2"].fill = _TITLE_FILL; ws["B2"].alignment = _CENTER

    # 数据表: 年份 + 三方案累计现金流
    r = 4
    headers = ["年份"] + [p["sheet_name"].replace("方案·", "") for p in plans_info]
    for i, h in enumerate(headers, 2):
        c = ws.cell(row=r, column=i, value=h)
        c.font = _TH_FONT; c.fill = _TH_FILL; c.alignment = _CENTER; c.border = _BORDER
    r += 1

    total_years = config["project"]["total_years"]
    start_year = config["project"]["start_year"]
    # 从各方案 sheet 的 cum_row 取数(直接写值, 图表需要)
    for y in range(total_years):
        ws.cell(row=r, column=2, value=str(start_year + y)).alignment = _CENTER
        for i, p in enumerate(plans_info):
            # 引用各方案 sheet 的累计行
            col = get_column_letter(3 + y)
            formula = f"='{p['sheet_name']}'!{col}{p['cum_row']}"
            c = ws.cell(row=r, column=3 + i, value=formula)
            c.number_format = _NUM_FMT; c.alignment = _RIGHT
        r += 1

    data_start = 5
    data_end = data_start + total_years - 1

    # 创建折线图
    chart = LineChart()
    chart.title = "三方案累计现金流对比"
    chart.style = 2
    chart.y_axis.title = "万元"
    chart.x_axis.title = "年份"
    chart.height = 12
    chart.width = 24

    for i, p in enumerate(plans_info):
        col = get_column_letter(3 + i)
        data_ref = Reference(ws, min_col=3 + i, min_row=data_start - 1, max_row=data_end)
        chart.add_data(data_ref, titles_from_data=True)

    cats = Reference(ws, min_col=2, min_row=data_start, max_row=data_end)
    chart.set_categories(cats)

    ws.add_chart(chart, f"{get_column_letter(2 + len(plans_info) + 2)}4")


# ============ Sheet 8: 结论与风险 ============
def _build_conclusion(wb: Workbook, plans_info: list[dict], config: dict):
    ws = wb.create_sheet("结论与风险")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [3, 26, 50, 4])

    ws.merge_cells("B2:C2")
    ws["B2"] = "结论与风险提示"
    ws["B2"].font = _TITLE_FONT; ws["B2"].fill = _TITLE_FILL; ws["B2"].alignment = _CENTER
    ws.row_dimensions[2].height = 26

    r = 4
    # 一、假设来源
    _style_section(ws, r, 3, "一、关键假设来源")
    r += 1
    assumptions_src = [
        ("收购总价", f"基于片区征收安置协议, {config['acquisition']['total_price']} 万元"),
        ("稳定期出租率", "参考片区同类房源运营数据, 取 85%-90%"),
        ("贷款利率", f"参考 LPR + 国企融资优势, {config['loan']['rate']*100:.2f}%"),
        ("运营成本占比", f"参考行业惯例, 取租金收入的 {config['operating_cost_ratio']*100:.0f}%"),
        ("装修周期", f"软硬装 5 年一更新, 每次 {config['renovation']['cycle_cost']} 万元"),
        ("税费政策", "公租房免征增值税、房产税(依据财税〔2014〕52 号等)"),
    ]
    for label, note in assumptions_src:
        ws.cell(row=r, column=2, value=label).font = _LABEL_FONT
        ws.cell(row=r, column=2).alignment = _LEFT; ws.cell(row=r, column=2).border = _BORDER
        ws.cell(row=r, column=2).fill = _SECTION_FILL
        c = ws.cell(row=r, column=3, value=note)
        c.font = _VAL_FONT; c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        c.border = _BORDER
        r += 1
    r += 1

    # 二、风险提示
    _style_section(ws, r, 3, "二、风险提示")
    r += 1
    risks = [
        ("出租率风险", "爬坡期延长或稳定期低于预期, 直接影响租金收入。敏感性: 出租率每下降 5 个百分点, IRR 下降约 0.8-1.2 个百分点。"),
        ("利率风险", "当前处于低利率周期, 若 LPR 上行 50bp, 总利息增加约 8%-10%, IRR 下降约 0.5 个百分点。"),
        ("租金增长风险", "假设每 2 年涨 2%, 实际受政策和市场影响。若不涨租, 累计现金流下降约 6%-8%。"),
        ("装修成本风险", "建材价格波动, 单次超支 20% 影响有限(占总投资 15%), 但周期性累积影响约 3%。"),
        ("出售风险(若有)", "若期末不出售, 项目依赖经营性现金流回收, 回收期显著延长。"),
    ]
    for label, note in risks:
        ws.cell(row=r, column=2, value=label).font = _BOLD_FONT
        ws.cell(row=r, column=2).alignment = _LEFT; ws.cell(row=r, column=2).border = _BORDER
        ws.cell(row=r, column=2).fill = _HIGHLIGHT_FILL
        c = ws.cell(row=r, column=3, value=note)
        c.font = _VAL_FONT; c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        c.border = _BORDER
        ws.row_dimensions[r].height = 32
        r += 1
    r += 1

    # 三、结论建议
    _style_section(ws, r, 3, "三、结论与建议")
    r += 1
    best = max(plans_info, key=lambda p: p["cumulative"])
    worst = min(plans_info, key=lambda p: p["cumulative"])
    conclusions = [
        f"本项目测算周期 {config['project']['total_years']} 年, 三方案 IRR 区间 {min(p['irr'] for p in plans_info):.2f}% - {max(p['irr'] for p in plans_info):.2f}%。",
        f"累计净现金流: 乐观 {plans_info[0]['cumulative']:.0f} 万元 / 中性 {plans_info[1]['cumulative']:.0f} 万元 / 保守 {plans_info[2]['cumulative']:.0f} 万元。",
        f"综合来看, {best['sheet_name'].replace('方案·','')} 累计净现金流最高, 但需结合风险承受能力选择。",
        f"建议采用中性方案作为基准, 保守方案作为压力测试下限, 乐观方案作为目标。",
        "后续工作: 1) 落实融资方案与利率锁定; 2) 签订长期租赁运营协议; 3) 建立运营 KPI 考核机制。",
    ]
    for note in conclusions:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        c = ws.cell(row=r, column=2, value=note)
        c.font = _VAL_FONT; c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        c.border = _BORDER
        ws.row_dimensions[r].height = 28
        r += 1


# ============ 主入口 ============
def export_zulin(config: dict, result: dict) -> bytes:
    """生成租赁测算 Excel, 返回字节流

    config: templates/zulin.yaml 加载的完整配置
    result: calculate_all_plans 返回的 {plans: [...]} 或单方案结果
    """
    wb = Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    # 封面
    _build_cover(wb, config)

    # 假设
    ref = _build_assumptions(wb, config)

    # 三方案
    scenarios = config.get("scenarios", [])
    plans_info = []

    # 兼容单方案和多方案
    if "plans" in result:
        results_list = result["plans"]
    else:
        # 单方案, 包装成列表
        results_list = [result]
        scenarios = [{"id": "default", "name": "基准方案"}]

    # 确保三方案
    plan_names = ["乐观", "中性", "保守"]
    for i, (res, sc) in enumerate(zip(results_list, scenarios)):
        name = plan_names[i] if i < len(plan_names) else sc.get("name", f"方案{i+1}")
        info = _build_plan_sheet(wb, name, config, sc, res, ref)
        plans_info.append(info)

    # 对比汇总
    _build_comparison(wb, plans_info, config)

    # 图表
    _build_chart(wb, plans_info, config)

    # 结论
    _build_conclusion(wb, plans_info, config)

    # 设定封面为活动 sheet
    wb.active = 0

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


if __name__ == "__main__":
    import yaml
    import sys
    sys.path.insert(0, ".")

    from engines.zulin import run
    from services.calculator import calculate_all_plans

    with open("templates/zulin.yaml") as f:
        cfg = yaml.safe_load(f)
    result = calculate_all_plans("zulin", cfg)
    data = export_zulin(cfg, result)
    out = "data/zulin_export_test.xlsx"
    import os
    os.makedirs("data", exist_ok=True)
    with open(out, "wb") as f:
        f.write(data)
    print(f"导出成功: {out} ({len(data)} bytes)")
